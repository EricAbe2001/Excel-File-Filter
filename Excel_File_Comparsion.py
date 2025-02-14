import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
from datetime import datetime

def compare_excel_files(file_a_path, file_b_path, output_file_path):
        """
    Compares two Excel files (A and B) and creates a new Excel file 
    highlighting the differences.  Only rows containing changes are included 
    in the output.

    values:
        file_a_path (str): Path to the first Excel file (file A).
        file_b_path (str): Path to the second Excel file (file B).
        output_file_path (str): Path to save the output Excel file.
    """
    try:
        # Read Excel files into pandas DataFrames
        df_a = pd.read_excel(file_a_path, header=None, engine='openpyxl')
        df_b = pd.read_excel(file_b_path, header=None, engine='openpyxl')

        # Handle empty files
        if df_a.empty or df_b.empty:
            print("Warning: One or both Excel files are empty.")
            wb = openpyxl.Workbook()
            ws = wb.active
            wb.save(output_file_path)
            return
            
        # Get maximum rows and columns for consistent comparison
        max_rows = max(df_a.shape[0], df_b.shape[0])
        max_cols = max(df_a.shape[1], df_b.shape[1])

        # Reindex and fill missing values to handle different dimensions
        df_a = df_a.reindex(index=range(max_rows), columns=range(max_cols)).fillna("")
        df_b = df_b.reindex(index=range(max_rows), columns=range(max_cols)).fillna("")

        # Create a new Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        output_row = 1 # Initialize output row counter

        # Iterate through rows and columns to compare cells
        for row in range(max_rows):
            row_has_difference = False # Flag for tracking changes in a row 
            for col in range(max_cols):
                # Check if cell indices are within the DataFrame bounds
                if row < df_a.shape[0] and col < df_a.shape[1] and row < df_b.shape[0] and col < df_b.shape[1]:
                    cell_a_value = str(df_a.iloc[row, col])
                    cell_b_value = str(df_b.iloc[row, col])

                    # Check for differences
                    if cell_a_value != cell_b_value:
                        row_has_difference = True  

             # If the row has difference, write it to the output Excel file
            if row_has_difference: 
                for col in range(max_cols): 
                    if row < df_a.shape[0] and col < df_a.shape[1] and row < df_b.shape[0] and col < df_b.shape[1]:
                        cell_a_value = str(df_a.iloc[row, col])
                        cell_b_value = str(df_b.iloc[row, col])
                        ws.cell(row=output_row, column=col + 1, value=cell_b_value) #write file b
                        if cell_a_value != cell_b_value:
                            red_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            ws.cell(row=output_row, column=col + 1).fill = red_fill
                    else:
                        ws.cell(row=output_row, column=col + 1, value="") 
                output_row += 1

       
        last_modified_file_b = os.path.getmtime(file_b_path)  # Get timestamp
        last_modified_datetime_file_b = datetime.fromtimestamp(last_modified_file_b)

         # Metadata tracking (last modified time and count)
        metadata_file = "comparison_metadata.txt"


         # Check if metadata file exists
        if os.path.exists(metadata_file):
            with open(metadata_file, "r", encoding="utf-8") as f:
                last_modified_str = f.readline().strip()
                try:
                    last_modified = datetime.strptime(last_modified_str, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    last_modified = datetime.now()
                modification_count = int(f.readline().strip())
                last_modified_file_b_stored = f.readline().strip() # read file b last modified
                try:
                    last_modified_file_b_stored = datetime.strptime(last_modified_file_b_stored, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    last_modified_file_b_stored = datetime.now()
        else:
            last_modified = datetime.now()
            modification_count = 0
            last_modified_file_b_stored = last_modified_datetime_file_b

        last_modified = datetime.now()
        modification_count += 1

        with open(metadata_file, "w", encoding="utf-8") as f:
            f.write(last_modified.strftime("%Y-%m-%d %H:%M:%S") + "\n")
            f.write(str(modification_count) + "\n")
            f.write(last_modified_datetime_file_b.strftime("%Y-%m-%d %H:%M:%S") + "\n") # write last modified file b

        wb.save(output_file_path)
        print(f"Comparison complete. Differences highlighted in {output_file_path}")
        print(f"Last modified: {last_modified.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Modification count: {modification_count}")
        print(f"Last modified of File B: {last_modified_datetime_file_b.strftime('%Y-%m-%d %H:%M:%S')}") # print to console

    except FileNotFoundError:
        print(f"Error: One or both files not found. Paths: {file_a_path}, {file_b_path}")
    except Exception as e:
        print(f"An error occurred: {e}")

file_a = "file_a.xlsx"
file_b = "file_b.xlsx"
output_file = "comparison_output.xlsx" 

compare_excel_files(file_a, file_b, output_file)
